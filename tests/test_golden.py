from __future__ import annotations

import warnings
from pathlib import Path

import polars as pl
import pytest

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
    module="openpyxl.worksheet._reader",
)

from openpyxl import load_workbook

from dpm._types import DeltaResult
from dpm.delta import compare_versions
from dpm.delta_schema import DELTA_STRUCTURE_COLS, STRUCTURE_XLSX
from dpm.excel import generate_delta_workbook
from dpm.workflows import _load_dataset, detect_version, load_perimeters, run_ingest

_SKIP_SHEETS = {"summary", "metrics", "dimensions"}
# The workbook writes columns in STRUCTURE_XLSX order; read them back positionally
# into the canonical snake_case names so both sides compare in the same schema.
_XLSX_KEYS = list(STRUCTURE_XLSX.keys())

DATA_DIR = Path(__file__).parent.parent / "data" / "input" / "public"
OLD_XLSX = (
    DATA_DIR
    / "EIOPA_Solvency_II_DPM_Annotated_Templates_2.8.2_Hotfix_table_group_arrangement.xlsx"
)
NEW_XLSX = (
    DATA_DIR
    / "EIOPA_Solvency_II_DPM_Annotated_Templates_2.10.0_table_group_arrangement.xlsx"
)

GOLDEN_DIR = Path(__file__).parent / "golden"
GOLDEN_FILE = GOLDEN_DIR / "delta.xlsx"

_SORT_COLS = [
    "perimeter",
    "template_code",
    "subtemplate_code",
    "row_code",
    "column_code",
    "status",
    "change_type",
]


def _run_delta(db_dir: Path) -> DeltaResult:
    # Real production path: ingest both workbooks to DuckDB, then delta from the DBs.
    old_version = detect_version(OLD_XLSX)
    new_version = detect_version(NEW_XLSX)
    assert old_version and new_version
    run_ingest(OLD_XLSX, old_version, db_dir)
    run_ingest(NEW_XLSX, new_version, db_dir)

    old_db = db_dir / f"{old_version}.duckdb"
    new_db = db_dir / f"{new_version}.duckdb"
    selected = {p.lower() for p in load_perimeters(old_db, new_db)}
    old_ds = _load_dataset(old_db, selected)
    new_ds = _load_dataset(new_db, selected)
    return compare_versions(old_ds, new_ds)


def _read_delta_from_xlsx(path: Path) -> pl.DataFrame:
    """Read all perimeter sheets from a delta workbook and concat into one DataFrame."""
    wb = load_workbook(path, read_only=True, data_only=True)
    frames: list[pl.DataFrame] = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in _SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        # first row is header, rest is data; read positionally into canonical names
        data = {
            col: [str(r[i]) if r[i] is not None else "" for r in rows[1:]]
            for i, col in enumerate(_XLSX_KEYS)
        }
        frames.append(pl.DataFrame(data))
    wb.close()
    return (
        pl.concat(frames, how="vertical")
        if frames
        else pl.DataFrame(schema={col: pl.String for col in DELTA_STRUCTURE_COLS})
    )


def _sorted(df: pl.DataFrame) -> pl.DataFrame:
    return df.sort(_SORT_COLS)


@pytest.mark.slow
def test_delta_golden(request: pytest.FixtureRequest, tmp_path: Path) -> None:
    actual = _run_delta(tmp_path)

    if request.config.getoption("--update-golden"):
        GOLDEN_DIR.mkdir(exist_ok=True)
        generate_delta_workbook(actual, GOLDEN_FILE)
        pytest.skip(f"Golden file updated: {GOLDEN_FILE}")

    if not GOLDEN_FILE.exists():
        pytest.fail(
            f"Golden file not found: {GOLDEN_FILE}\n"
            "Run with --update-golden to generate it."
        )

    expected = _sorted(_read_delta_from_xlsx(GOLDEN_FILE))
    actual_sorted = _sorted(
        actual.structure.select(
            [pl.col(c).cast(pl.String).fill_null("") for c in DELTA_STRUCTURE_COLS]
        )
    )

    if not actual_sorted.equals(expected):
        lines: list[str] = []

        # rows in actual not in golden (added)
        only_actual = actual_sorted.join(expected, on=_SORT_COLS, how="anti")
        # rows in golden not in actual (removed)
        only_expected = expected.join(actual_sorted, on=_SORT_COLS, how="anti")

        n_added = only_actual.height
        n_removed = only_expected.height

        if n_added:
            lines.append(
                f"\n--- {n_added} row(s) present in actual but not in golden ---"
            )
            lines.append(str(only_actual.head(20)))
            if n_added > 20:
                lines.append(f"  ... and {n_added - 20} more")

        if n_removed:
            lines.append(
                f"\n--- {n_removed} row(s) present in golden but not in actual ---"
            )
            lines.append(str(only_expected.head(20)))
            if n_removed > 20:
                lines.append(f"  ... and {n_removed - 20} more")

        if not n_added and not n_removed:
            # same keys, different non-key column values
            diff_mask = actual_sorted != expected
            changed_rows = diff_mask.select(pl.any_horizontal(pl.all())).to_series()
            changed_cols = [c for c in DELTA_STRUCTURE_COLS if diff_mask[c].any()]
            lines.append(
                f"\n--- {changed_rows.sum()} row(s) with changed values in: {changed_cols} ---"
            )
            lines.append(str(actual_sorted.filter(changed_rows).head(20)))

        lines.append("\nRun with --update-golden to accept new output as golden.")
        pytest.fail("\n".join(lines))
