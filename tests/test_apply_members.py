"""Unit tests for dimensional-context patching (context re-pointing) in apply-delta."""

from __future__ import annotations

import logging

import polars as pl

from dpm.xbrl import apply_context_remap, build_context_remap, parse_contexts


def _delta(rows: list[dict[str, str]]) -> pl.DataFrame:
    cols = ["qname_old", "status", "dimensions_old", "dimensions_new"]
    return pl.DataFrame(
        [{c: r.get(c, "") for c in cols} for r in rows],
        schema={c: pl.String for c in cols},
    )


def _s(spec: str) -> frozenset[tuple[str, str]]:
    return frozenset(
        (d.split("=")[0], d.split("=")[1]) for d in spec.split(";") if "=" in d
    )


def _remap(*entries: tuple[str, str, str]) -> dict:
    """Build a ContextRemap from ``(qname, old_spec, new_spec)`` triples."""
    out: dict = {}
    for qname, old, new in entries:
        out.setdefault(qname, []).append((_s(old), _s(new)))
    for changes in out.values():
        changes.sort(key=lambda pair: len(pair[0]), reverse=True)
    return out


# ── build_context_remap ─────────────────────────────────────────────────────


def test_remap_keyed_on_qname_and_full_set():
    delta = _delta(
        [
            {
                "qname_old": "s2md_met:mi1",
                "status": "Modified",
                "dimensions_old": "s2c_dim:VG=s2c_AM:x80;s2c_dim:RT=s2c_RT:x85",
                "dimensions_new": "s2c_dim:VG=s2c_AM:x81;s2c_dim:RT=s2c_RT:x85",
            }
        ]
    )
    remap = build_context_remap(delta)
    assert remap == {
        "s2md_met:mi1": [
            (
                _s("s2c_dim:VG=s2c_AM:x80;s2c_dim:RT=s2c_RT:x85"),
                _s("s2c_dim:VG=s2c_AM:x81;s2c_dim:RT=s2c_RT:x85"),
            )
        ]
    }


def test_same_member_different_context_are_distinct_changes():
    # The exact ambiguity the single-member approach got wrong: DV=x11 maps to two
    # different targets depending on the rest of the cell — kept as two distinct cells.
    delta = _delta(
        [
            {
                "qname_old": "s2md_met:mi1",
                "status": "Modified",
                "dimensions_old": "s2c_dim:DV=s2c_RT:x11;s2c_dim:RT=s2c_RT:x85",
                "dimensions_new": "s2c_dim:DV=s2c_RT:x0;s2c_dim:RT=s2c_RT:x85",
            },
            {
                "qname_old": "s2md_met:mi1",
                "status": "Modified",
                "dimensions_old": "s2c_dim:DV=s2c_RT:x11;s2c_dim:RT=s2c_RT:x232",
                "dimensions_new": "s2c_dim:DV=s2c_RT:x352;s2c_dim:RT=s2c_RT:x232",
            },
        ]
    )
    remap = build_context_remap(delta)
    assert list(remap) == ["s2md_met:mi1"]
    assert len(remap["s2md_met:mi1"]) == 2  # no conflict — the full context disambiguates


def test_whole_dimension_add_is_captured_as_full_set_replacement():
    delta = _delta(
        [
            {
                "qname_old": "s2md_met:mi1",
                "status": "Modified",
                "dimensions_old": "s2c_dim:VG=s2c_AM:x80",
                "dimensions_new": "s2c_dim:VG=s2c_AM:x80;s2c_dim:RT=s2c_RT:x85",
            }
        ]
    )
    remap = build_context_remap(delta)
    (changes,) = remap.values()
    assert changes[0][1] == _s("s2c_dim:VG=s2c_AM:x80;s2c_dim:RT=s2c_RT:x85")


def test_unchanged_and_non_modified_ignored():
    delta = _delta(
        [
            {"qname_old": "s2md_met:mi1", "status": "Modified",
             "dimensions_old": "s2c_dim:VG=s2c_AM:x80", "dimensions_new": "s2c_dim:VG=s2c_AM:x80"},
            {"qname_old": "s2md_met:mi2", "status": "Deleted",
             "dimensions_old": "s2c_dim:VG=s2c_AM:x80", "dimensions_new": ""},
        ]
    )
    assert build_context_remap(delta) == {}


def test_conflicting_key_dropped_and_logged(caplog):
    delta = _delta(
        [
            {"qname_old": "s2md_met:mi1", "status": "Modified",
             "dimensions_old": "s2c_dim:VG=s2c_AM:x80", "dimensions_new": "s2c_dim:VG=s2c_AM:x81"},
            {"qname_old": "s2md_met:mi1", "status": "Modified",
             "dimensions_old": "s2c_dim:VG=s2c_AM:x80", "dimensions_new": "s2c_dim:VG=s2c_AM:x82"},
        ]
    )
    with caplog.at_level(logging.WARNING):
        remap = build_context_remap(delta)
    assert remap == {}  # ambiguous key removed entirely
    assert "Conflicting context remap" in caplog.text


# ── apply_context_remap ─────────────────────────────────────────────────────


def _ctx(cid: str, members: str, typed: str = "") -> bytes:
    explicit = "".join(
        f'<xbrldi:explicitMember dimension="{d}">{m}</xbrldi:explicitMember>'
        for d, m in (p.split("=") for p in members.split(";") if p)
    )
    return (
        f'<xbrli:context id="{cid}">'
        f"<xbrli:entity><xbrli:identifier>LEI</xbrli:identifier></xbrli:entity>"
        f"<xbrli:period><xbrli:instant>2025-12-31</xbrli:instant></xbrli:period>"
        f"<xbrli:scenario>{typed}{explicit}</xbrli:scenario>"
        f"</xbrli:context>"
    ).encode()


def _fact(qname: str, ref: str, value: str = "1") -> bytes:
    return f'<s2md_met:{qname} contextRef="{ref}">{value}</s2md_met:{qname}>'.encode()


def _doc(*bodies: bytes) -> bytes:
    return b"<xbrli:xbrl>" + b"".join(bodies) + b"</xbrli:xbrl>"


def test_single_qname_context_repointed_to_new_context():
    xml = _doc(_ctx("c1", "s2c_dim:VG=s2c_AM:x80"), _fact("mi1", "c1"))
    remap = _remap(("s2md_met:mi1", "s2c_dim:VG=s2c_AM:x80", "s2c_dim:VG=s2c_AM:x81"))
    res = apply_context_remap(xml, remap)
    out = res.xml
    assert (res.repointed_facts, res.new_contexts) == (1, 1)
    ctxs = parse_contexts(out)
    # the fact now points at a context carrying the new member set
    new_id = out.split(b'contextRef="')[1].split(b'"')[0].decode()
    assert ctxs[new_id].members == _s("s2c_dim:VG=s2c_AM:x81")
    assert new_id != "c1"
    # detail frames describe the change for the debug workbook
    detail = res.repoint_details.to_dicts()[0]
    assert detail["qname"] == "s2md_met:mi1"
    assert detail["context_old"] == "c1" and detail["context_new"] == new_id
    assert detail["dimensions_old"] == "s2c_dim:VG=s2c_AM:x80"
    assert detail["dimensions_new"] == "s2c_dim:VG=s2c_AM:x81"
    assert res.new_context_details.to_dicts()[0]["cloned_from"] == "c1"


def test_shared_context_only_matching_qname_moves():
    # c1 is shared by mi1 (changes) and mi2 (unchanged): mi2 must stay on c1.
    xml = _doc(
        _ctx("c1", "s2c_dim:VG=s2c_AM:x80"),
        _fact("mi1", "c1"),
        _fact("mi2", "c1"),
    )
    remap = _remap(("s2md_met:mi1", "s2c_dim:VG=s2c_AM:x80", "s2c_dim:VG=s2c_AM:x81"))
    res = apply_context_remap(xml, remap)
    out = res.xml
    assert (res.repointed_facts, res.new_contexts) == (1, 1)
    mi1 = out.split(b"mi1 contextRef=\"")[1].split(b'"')[0]
    mi2 = out.split(b"mi2 contextRef=\"")[1].split(b'"')[0]
    assert mi2 == b"c1"          # untouched fact keeps original context
    assert mi1 != b"c1"          # changed fact moved
    assert parse_contexts(out)[mi1.decode()].members == _s("s2c_dim:VG=s2c_AM:x81")


def test_typed_member_preserved_on_clone():
    typed = '<xbrldi:typedMember dimension="s2c_dim:PX"><s2c_typ:ID>1</s2c_typ:ID></xbrldi:typedMember>'
    xml = _doc(_ctx("c1", "s2c_dim:VG=s2c_AM:x80", typed=typed), _fact("mi1", "c1"))
    remap = _remap(("s2md_met:mi1", "s2c_dim:VG=s2c_AM:x80", "s2c_dim:VG=s2c_AM:x81"))
    out = apply_context_remap(xml, remap).xml
    new_id = out.split(b'contextRef="')[1].split(b'"')[0].decode()
    block = [m for c, m in _blocks(out).items() if c == new_id][0]
    assert b"typedMember" in block and b"s2c_typ:ID" in block


def test_existing_target_context_reused_no_duplicate():
    # c2 already has the target member set — the fact should re-point to c2, not clone.
    xml = _doc(
        _ctx("c1", "s2c_dim:VG=s2c_AM:x80"),
        _ctx("c2", "s2c_dim:VG=s2c_AM:x81"),
        _fact("mi1", "c1"),
    )
    remap = _remap(("s2md_met:mi1", "s2c_dim:VG=s2c_AM:x80", "s2c_dim:VG=s2c_AM:x81"))
    res = apply_context_remap(xml, remap)
    assert (res.repointed_facts, res.new_contexts) == (1, 0)  # reused, nothing created
    assert res.xml.split(b'mi1 contextRef="')[1].split(b'"')[0] == b"c2"


def test_no_remap_is_identity():
    xml = _doc(_ctx("c1", "s2c_dim:VG=s2c_AM:x80"), _fact("mi1", "c1"))
    res = apply_context_remap(xml, {})
    assert res.xml == xml and res.repointed_facts == 0 and res.new_contexts == 0


def test_subset_match_rewrites_only_named_dimension():
    # The delta records only the varied dimension (VI); the instance context also carries
    # the implicit VG=x80. Subset matching must apply VI x64→x116 and KEEP VG=x80.
    xml = _doc(
        _ctx("c1", "s2c_dim:VI=s2c_VM:x64;s2c_dim:VG=s2c_AM:x80"),
        _fact("mi363", "c1"),
    )
    remap = _remap(("s2md_met:mi363", "s2c_dim:VI=s2c_VM:x64", "s2c_dim:VI=s2c_VM:x116"))
    res = apply_context_remap(xml, remap)
    assert res.repointed_facts == 1
    new_id = res.xml.split(b'contextRef="')[1].split(b'"')[0].decode()
    assert parse_contexts(res.xml)[new_id].members == _s(
        "s2c_dim:VI=s2c_VM:x116;s2c_dim:VG=s2c_AM:x80"
    )


def test_most_specific_change_wins():
    # Two changes for the same qname both subset the context; the more specific (larger
    # old_set) must win: {VI=x64, RT=x9} → target, not the generic {VI=x64} → other.
    xml = _doc(
        _ctx("c1", "s2c_dim:VI=s2c_VM:x64;s2c_dim:RT=s2c_RT:x9;s2c_dim:VG=s2c_AM:x80"),
        _fact("mi1", "c1"),
    )
    remap = _remap(
        ("s2md_met:mi1", "s2c_dim:VI=s2c_VM:x64", "s2c_dim:VI=s2c_VM:x999"),
        (
            "s2md_met:mi1",
            "s2c_dim:VI=s2c_VM:x64;s2c_dim:RT=s2c_RT:x9",
            "s2c_dim:VI=s2c_VM:x116;s2c_dim:RT=s2c_RT:x9",
        ),
    )
    res = apply_context_remap(xml, remap)
    new_id = res.xml.split(b'contextRef="')[1].split(b'"')[0].decode()
    members = parse_contexts(res.xml)[new_id].members
    assert ("s2c_dim:VI", "s2c_VM:x116") in members  # specific change applied
    assert ("s2c_dim:VI", "s2c_VM:x999") not in members  # generic one did not win


def test_partial_context_metric_split():
    # A shared context where only SOME metrics change VI: the changed ones move to one
    # new context; the unchanged metric stays. Proves per-metric splitting on a superset.
    xml = _doc(
        _ctx("c1", "s2c_dim:VI=s2c_VM:x64;s2c_dim:VG=s2c_AM:x80"),
        _fact("mi363", "c1"),
        _fact("mi490", "c1"),
        _fact("pi548", "c1"),
    )
    remap = _remap(
        ("s2md_met:mi363", "s2c_dim:VI=s2c_VM:x64", "s2c_dim:VI=s2c_VM:x116"),
        ("s2md_met:mi490", "s2c_dim:VI=s2c_VM:x64", "s2c_dim:VI=s2c_VM:x116"),
    )
    res = apply_context_remap(xml, remap)
    assert (res.repointed_facts, res.new_contexts) == (2, 1)  # two facts, one shared clone
    refs = {
        q: res.xml.split(f"{q} contextRef=\"".encode())[1].split(b'"')[0]
        for q in ("mi363", "mi490", "pi548")
    }
    assert refs["mi363"] == refs["mi490"] != b"c1"  # both changed → same new context
    assert refs["pi548"] == b"c1"  # unchanged metric stays put


def _blocks(xml: bytes) -> dict[str, bytes]:
    from dpm.xbrl import CONTEXT_RE

    return {m.group("id").decode(): m.group(0) for m in CONTEXT_RE.finditer(xml)}


# ── debug workbook ──────────────────────────────────────────────────────────


def test_debug_workbook_sheets_and_nesting(tmp_path):
    import openpyxl

    from dpm._types import ApplyStats
    from dpm.excel import generate_apply_debug_workbook

    stats = ApplyStats(
        perimeter="ars", facts_before=10, facts_after=9, deleted_facts=1,
        renamed_facts=1, deleted_qnames=1, modified_qnames=1,
        repointed_facts=2, new_contexts=1,
    )
    deleted = pl.DataFrame(
        {"qname": ["s2md_met:mi1"], "qname_new": ["s2md_met:mi1"],
         "context_ref": ["c1"], "dimensions": ["s2c_dim:VG=s2c_AM:x80"], "value": ["5"]}
    )
    renamed = pl.DataFrame(
        {"qname": ["s2md_met:mi2"], "qname_new": ["s2md_met:mi9"],
         "context_ref": ["c2"], "dimensions": [""], "value": ["7"]}
    )
    repointed = pl.DataFrame(
        {"qname": ["s2md_met:mi3", "s2md_met:mi3"], "context_old": ["c3", "c4"],
         "context_new": ["cdpm1", "cdpm1"],
         "dimensions_old": ["s2c_dim:RT=s2c_RT:x148"] * 2,
         "dimensions_new": ["s2c_dim:RT=s2c_RT:x355"] * 2, "value": ["1", "2"]}
    )
    new_contexts = pl.DataFrame(
        {"context_id": ["cdpm1"], "dimensions": ["s2c_dim:RT=s2c_RT:x355"], "cloned_from": ["c3"]}
    )

    out = tmp_path / "debug.xlsx"
    generate_apply_debug_workbook(
        out, stats=stats, deleted=deleted, renamed=renamed,
        repointed=repointed, new_contexts=new_contexts,
    )

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == [
        "Summary", "Deleted metrics", "Renamed metrics", "Re-pointed cells", "New contexts",
    ]
    # Deleted / Renamed are flat: one row per metric (no per-fact detail, since a
    # rename/delete leaves each fact's context, dimensions and value unchanged).
    dl = wb["Deleted metrics"]
    assert [dl["A2"].value, dl["B2"].value] == ["s2md_met:mi1", 1] and dl.max_row == 2
    rn = wb["Renamed metrics"]
    assert [rn["A2"].value, rn["B2"].value] == ["s2md_met:mi2 → s2md_met:mi9", 1]
    assert rn.max_row == 2  # single metric, no nested fact rows
    # Re-pointed keeps nesting: one metric parent (mi3) + two nested fact rows (level 1).
    ws = wb["Re-pointed cells"]
    assert ws["A2"].value == "s2md_met:mi3" and ws["B2"].value == "2 facts"
    assert ws.row_dimensions[3].outline_level == 1
    assert ws.row_dimensions[4].outline_level == 1
    assert "cdpm1" in str(ws["B3"].value)  # child shows context_old → context_new
    # Summary carries the stats.
    assert wb["Summary"]["A2"].value == "ars"
