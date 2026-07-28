#!/usr/bin/env python3
from __future__ import annotations

import argparse
import logging
import re
import sys
import warnings
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple
from tqdm import tqdm
import pandas as pd

try:
    import polars as pl
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: polars. Install with: pip install polars"
    ) from exc

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
    module="openpyxl.worksheet._reader",
)

LOG = logging.getLogger("dpm_delta")

ROW_RE = re.compile(r"\b[A-Z]{0,3}R\d{3,6}\b", re.I)
COL_RE = re.compile(r"\b[A-Z]{0,3}C\d{3,6}\b", re.I)
CODE_RE = re.compile(r"\b[A-Z]{1,3}\.(?:[0-9]{2}\.){1,6}[0-9]{2}\b", re.I)
QNAME_RE = re.compile(
    r"(^\{[^}]+\}.+)|(^[A-Za-z_][\w.-]*:[A-Za-z_][\w.-]*$)|(^[A-Za-z_][\w.-]*\.[A-Za-z_][\w.-]*(?:\.[A-Za-z_][\w.-]*)+$)"
)
QNAME_TOKEN_RE = re.compile(r"\b(?:s2md_met|[A-Za-z_][\w.-]*):[A-Za-z_][\w.-]*\b")
METRIC_LABEL_RE = re.compile(r"Metric:\s*(.*?)(?:\)\s*(?:\[|$)|$)")
GENERIC_TOC_CODES = {"T99", "T.99", "T.99.99", "TOC", "TABLE OF CONTENTS"}

METRIC_COLS = [
    "perimeter",
    "template_code",
    "subtemplate_code",
    "row_code",
    "column_code",
    "qname",
    "metric_label",
]
KEY_COLS = ["perimeter", "template_code", "subtemplate_code", "row_code", "column_code"]
IDENTITY_COLS = ["template_code", "subtemplate_code", "row_code", "column_code"]
DELTA_COLS = [
    "Perimeter",
    "TemplateCode",
    "SubTemplateCode",
    "RowCode",
    "ColumnCode",
    "QNameOld",
    "QNameNew",
    "MetricLabelOld",
    "MetricLabelNew",
    "Status",
    "ChangeType",
    "Comments",
]

ADDED_FILL = PatternFill("solid", fgColor="C6EFCE")
DELETED_FILL = PatternFill("solid", fgColor="FFC7CE")
MODIFIED_FILL = PatternFill("solid", fgColor="FCE4D6")
KEPT_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


@dataclass(frozen=True)
class TocEntry:
    perimeter: str
    template_code: str
    subtemplate_code: str
    sheet_name: Optional[str] = None


@dataclass(frozen=True)
class DpmDataset:
    metrics: pl.DataFrame
    entries: List[TocEntry]

    @property
    def templates(self) -> Set[Tuple[str, str]]:
        return {(e.perimeter, e.template_code) for e in self.entries}

    @property
    def subtemplates(self) -> Set[Tuple[str, str, str]]:
        return {
            (e.perimeter, e.template_code, e.subtemplate_code) for e in self.entries
        }


@dataclass
class WorkbookCache:
    path: Path
    wb: Any
    sheet_names: List[str]
    _matrix_cache: Dict[str, List[List[str]]]

    @classmethod
    def open(cls, path: Path) -> "WorkbookCache":
        wb = load_workbook(path, read_only=True, data_only=True)
        return cls(
            path=path,
            wb=wb,
            sheet_names=list(wb.sheetnames),
            _matrix_cache={},
        )

    def close(self) -> None:
        self.wb.close()

    def matrix(self, sheet_name: str) -> List[List[str]]:
        if sheet_name not in self._matrix_cache:
            ws = self.wb[sheet_name]

            self._matrix_cache[sheet_name] = [
                [norm(v) for v in row] for row in ws.iter_rows(values_only=True)
            ]

        return self._matrix_cache[sheet_name]


def empty_metric_df() -> pl.DataFrame:
    return pl.DataFrame(schema={col: pl.Utf8 for col in METRIC_COLS})


def empty_delta_df() -> pl.DataFrame:
    return pl.DataFrame(schema={col: pl.Utf8 for col in DELTA_COLS})


def norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ").replace("\r", " ").strip())



def metric_change_type_expr(row_col: str = "row_code") -> pl.Expr:
    return (
        pl.when(pl.col(row_col) != "")
        .then(pl.lit("MetricRow"))
        .otherwise(pl.lit("MetricColumn"))
    )



def first_match(pattern: re.Pattern[str], text: str) -> str:
    match = pattern.search(text or "")
    return match.group(0).upper() if match else ""


def extract_qname(text: str) -> str:
    match = QNAME_TOKEN_RE.search(norm(text))
    return match.group(0) if match else ""


def extract_metric_label(text: str) -> str:
    match = METRIC_LABEL_RE.search(norm(text))
    return match.group(1).strip() if match else ""


def is_qname(text: str) -> bool:
    text = norm(text)
    return bool(extract_qname(text) or QNAME_RE.search(text))


def discover_toc_sheet(workbook: WorkbookCache) -> str:
    names = workbook.sheet_names

    exact = next(
        (name for name in names if name == "Table of Contents"),
        None,
    )
    if exact:
        LOG.info("TOC sheet: %s", exact)
        return exact

    case_insensitive = next(
        (name for name in names if name.lower() == "table of contents"),
        None,
    )
    if case_insensitive:
        LOG.info("TOC sheet: %s", case_insensitive)
        return case_insensitive

    raise ValueError('Required TOC sheet "Table of Contents" not found')


def looks_like_perimeter(value: str) -> bool:
    value = norm(value).lower()
    if (
        not value
        or CODE_RE.search(value)
        or ROW_RE.search(value)
        or COL_RE.search(value)
    ):
        return False
    forbidden = {
        "perimeter",
        "scope",
        "template",
        "templates",
        "subtemplate",
        "sub-template",
        "sheet",
        "code",
        "name",
        "table",
        "contents",
        "toc",
        "description",
    }
    return bool(
        value not in forbidden
        and len(value) <= 12
        and re.fullmatch(r"[a-z][a-z0-9_-]{1,11}", value)
    )


def parent_template_code(subtemplate_code: str, current_template: str = "") -> str:
    sub = norm(subtemplate_code).upper()
    cur = norm(current_template).upper()
    if cur and sub.startswith(cur + "."):
        return cur
    parts = sub.split(".")
    return ".".join(parts[:-1]) if len(parts) > 3 else sub


def find_horizontal_header_row(rows: List[List[str]]) -> int:
    best_i, best = 0, -1
    for i, row in enumerate(rows[:80]):
        low = [c.lower() for c in row]
        score = 100 if any(c == "template code" for c in low) else 0
        score += 10 if any(c == "description" for c in low) else 0
        score += sum(5 for c in low if looks_like_perimeter(c))
        if score > best:
            best_i, best = i, score
    return best_i


def extract_toc(workbook: WorkbookCache) -> List[TocEntry]:
    toc_sheet = discover_toc_sheet(workbook)
    raw = workbook.matrix(toc_sheet)
    if not raw:
        raise ValueError("TOC sheet is empty")
    for i, row in enumerate(raw[:15], start=1):
        LOG.debug("TOC first lines %s: %s", i, row[:30])

    width = max(len(row) for row in raw)
    rows = [row + [""] * (width - len(row)) for row in raw]
    header_i = find_horizontal_header_row(rows)
    headers = rows[header_i]

    code_col = next(
        (
            i
            for i, h in enumerate(headers)
            if h.lower() in {"template code", "template", "code"}
        ),
        -1,
    )
    desc_col = next(
        (i for i, h in enumerate(headers) if h.lower() == "description"), -1
    )
    perimeter_cols = [
        (i, h.lower())
        for i, h in enumerate(headers)
        if i not in {code_col, desc_col} and looks_like_perimeter(h)
    ]
    if code_col < 0 or not perimeter_cols:
        raise ValueError(
            "Unable to parse 'Table of Contents': expected Template code, Description, then perimeter columns"
        )

    entries: Dict[Tuple[str, str, str], TocEntry] = {}
    current_template = ""
    for row in rows[header_i + 1 :]:
        row_code = first_match(CODE_RE, row[code_col]) if code_col < len(row) else ""
        if not row_code or row_code in GENERIC_TOC_CODES or row_code.startswith("T.99"):
            continue
        populated: List[Tuple[str, str]] = []
        for col, perimeter in perimeter_cols:
            cell_code = first_match(CODE_RE, row[col]) if col < len(row) else ""
            if (
                cell_code
                and cell_code not in GENERIC_TOC_CODES
                and not cell_code.startswith("T.99")
            ):
                populated.append((perimeter, cell_code))
        if not populated:
            current_template = row_code
            continue
        for perimeter, subtemplate in populated:
            template = parent_template_code(subtemplate, current_template)
            entries[(perimeter, template, subtemplate)] = TocEntry(
                perimeter, template, subtemplate
            )

    if not entries:
        raise ValueError(
            "TOC extraction returned no perimeter/template/subtemplate entries"
        )
    distribution = Counter(e.perimeter for e in entries.values())
    LOG.info(
        "TOC entries=%d perimeters=%s", len(entries), dict(sorted(distribution.items()))
    )
    return sorted(
        entries.values(),
        key=lambda e: (e.perimeter, e.template_code, e.subtemplate_code),
    )


def extract_perimeters(entries: Sequence[TocEntry]) -> Set[str]:
    return {e.perimeter for e in entries}


def extract_templates(entries: Sequence[TocEntry]) -> Set[Tuple[str, str]]:
    return {(e.perimeter, e.template_code) for e in entries}


def extract_subtemplates(entries: Sequence[TocEntry]) -> Set[Tuple[str, str, str]]:
    return {(e.perimeter, e.template_code, e.subtemplate_code) for e in entries}


def parse_text_selection(raw: str, perimeters: Sequence[str]) -> Set[str]:
    raw = raw.strip().lower()
    if not raw or raw == "all":
        return {p.lower() for p in perimeters}
    selected: Set[str] = set()
    by_name = {p.lower(): p.lower() for p in perimeters}
    for token in re.split(r"[,;\s]+", raw):
        if not token:
            continue
        if "-" in token and all(part.isdigit() for part in token.split("-", 1)):
            start, end = [int(part) for part in token.split("-", 1)]
            for idx in range(start, end + 1):
                if 1 <= idx <= len(perimeters):
                    selected.add(perimeters[idx - 1].lower())
        elif token.isdigit():
            idx = int(token)
            if 1 <= idx <= len(perimeters):
                selected.add(perimeters[idx - 1].lower())
        elif token in by_name:
            selected.add(by_name[token])
        else:
            raise ValueError(f"Unknown perimeter selection token: {token!r}")
    if not selected:
        raise ValueError("At least one perimeter must be selected")
    return selected


def select_perimeters_interactively(
    old_entries: Sequence[TocEntry], new_entries: Sequence[TocEntry]
) -> Set[str]:
    perimeters = sorted(
        extract_perimeters(old_entries) | extract_perimeters(new_entries)
    )
    if not perimeters:
        raise ValueError("No perimeters discovered in TOC")
    if sys.stdin.isatty():
        try:
            import questionary  # type: ignore
            from questionary import Choice  # type: ignore

            choices = [Choice(title=p, value=p, checked=True) for p in perimeters]
            selected = questionary.checkbox(
                "Select reporting perimeters:", choices=choices
            ).ask()
            if selected is None:
                raise KeyboardInterrupt("Perimeter selection cancelled")
            if not selected:
                raise ValueError("At least one perimeter must be selected")
            return {str(p).lower() for p in selected}
        except ImportError:
            pass
    print("\nDetected reporting perimeters:")
    for idx, perimeter in enumerate(perimeters, start=1):
        print(f"  [{idx:02d}] {perimeter}")
    if sys.stdin.isatty():
        print("\nInstall questionary for checkboxes: pip install questionary")
        print("Select by number/name/range, e.g. '1 3-5 ars qrs'. Press Enter for all.")
        return parse_text_selection(input("Perimeters> "), perimeters)
    LOG.warning(
        "Non-interactive terminal detected; selecting all perimeters. Use --no-interactive-perimeters to make this explicit."
    )
    return {p.lower() for p in perimeters}


def filter_entries_by_perimeter(
    entries: Sequence[TocEntry], selected: Set[str]
) -> List[TocEntry]:
    return [e for e in entries if e.perimeter.lower() in selected]


def candidate_sheets(entry: TocEntry, sheets: Sequence[str]) -> List[str]:
    by_lower = {sheet.lower(): sheet for sheet in sheets}
    exact = by_lower.get(entry.template_code.lower())
    if exact:
        return [exact]
    if entry.sheet_name and entry.sheet_name.lower() in by_lower:
        return [by_lower[entry.sheet_name.lower()]]
    raise KeyError(f"No worksheet named like template {entry.template_code!r}")


def first_non_empty_cell(row: Sequence[str]) -> Tuple[int, str]:
    for col, cell in enumerate(row):
        if norm(cell):
            return col, norm(cell)
    return -1, ""


def is_subtemplate_header_row(
    row: Sequence[str], subtemplate_code: Optional[str] = None
) -> bool:
    col, cell = first_non_empty_cell(row)
    if col != 0 or not cell:
        return False
    if subtemplate_code:
        return bool(re.match(rf"^{re.escape(subtemplate_code)}\s+-\s+", cell, re.I))
    return bool(
        re.match(r"^[A-Z]{1,3}\.(?:[0-9]{2}\.){2,7}[0-9]{2}\s+-\s+", cell, re.I)
    )


def subtemplate_window(
    matrix: List[List[str]], subtemplate_code: str
) -> List[List[str]]:
    anchors = [
        idx
        for idx, row in enumerate(matrix)
        if is_subtemplate_header_row(row, subtemplate_code)
    ]
    all_anchors = [
        idx for idx, row in enumerate(matrix) if is_subtemplate_header_row(row)
    ]
    if not anchors:
        LOG.debug(
            "Subtemplate header not found for %s; using full sheet window",
            subtemplate_code,
        )
        return matrix
    start = anchors[0]
    end = min((idx for idx in all_anchors if idx > start), default=len(matrix))
    return matrix[start:end]


def nearest_text(values: Iterable[str]) -> str:
    for value in values:
        text = norm(value)
        if (
            text
            and not ROW_RE.fullmatch(text)
            and not COL_RE.fullmatch(text)
            and not CODE_RE.fullmatch(text)
            and text.lower() != "metrics"
            and not is_qname(text)
        ):
            return text[:500]
    return ""


def first_coord(pattern: re.Pattern[str], value: str) -> str:
    return first_match(pattern, value)


def find_col_code_for_metric_col(
    matrix: List[List[str]], qrow: int, qcol: int, search_start: int, search_end: int
) -> str:
    candidates: List[Tuple[int, int, str]] = []
    for r in range(max(0, search_start), min(len(matrix), search_end + 1)):
        for c in range(max(0, qcol - 2), min(len(matrix[r]), qcol + 1)):
            code = first_coord(COL_RE, matrix[r][c])
            if code:
                candidates.append((abs(qrow - r) + abs(qcol - c), c, code))
    if candidates:
        return sorted(candidates, key=lambda item: (item[0], -item[1]))[0][2]
    all_codes: List[str] = []
    for r in range(max(0, search_start), min(len(matrix), search_end + 1)):
        for cell in matrix[r]:
            code = first_coord(COL_RE, cell)
            if code:
                all_codes.append(code)
    return all_codes[0] if len(set(all_codes)) == 1 else ""


def row_label_for(matrix: List[List[str]], row_idx: int, row_code_col: int) -> str:
    left = matrix[row_idx][:row_code_col]
    right = matrix[row_idx][row_code_col + 1 :]
    return nearest_text(reversed(left)) or nearest_text(right)


def extract_row_oriented_rows(
    matrix: List[List[str]],
    perimeter: str,
    template: str,
    subtemplate: str,
    metrics_row: int,
    metrics_col: int,
) -> List[Dict[str, str]]:
    out: Dict[Tuple[str, str], Dict[str, str]] = {}
    row_items: List[Tuple[int, int, str]] = []
    for r in range(metrics_row + 1, len(matrix)):
        for c, cell in enumerate(matrix[r]):
            row_code = first_coord(ROW_RE, cell)
            if row_code:
                row_items.append((r, c, row_code))
                break
    if not row_items:
        return []
    first_data_row = row_items[0][0]
    for r, row_code_col, row_code in row_items:
        label = row_label_for(matrix, r, row_code_col)
        for qcol, cell in enumerate(matrix[r]):
            if qcol < metrics_col and not is_qname(cell):
                continue
            if not is_qname(cell):
                continue
            col_code = find_col_code_for_metric_col(
                matrix, r, qcol, metrics_row, first_data_row
            )
            if not col_code:
                continue
            out.setdefault(
                (row_code, col_code),
                {
                    "perimeter": perimeter,
                    "template_code": template,
                    "subtemplate_code": subtemplate,
                    "row_code": row_code,
                    "column_code": col_code,
                    "qname": extract_qname(cell) or cell,
                    "metric_label": extract_metric_label(cell) or label,
                },
            )
    return list(out.values())


def extract_column_oriented_rows(
    matrix: List[List[str]],
    perimeter: str,
    template: str,
    subtemplate: str,
    metrics_row: int,
    metrics_col: int,
) -> List[Dict[str, str]]:
    out: Dict[Tuple[str, str], Dict[str, str]] = {}
    for qcol in range(metrics_col + 1, len(matrix[metrics_row])):
        qname = matrix[metrics_row][qcol]
        if not is_qname(qname):
            continue
        col_code = find_col_code_for_metric_col(
            matrix, metrics_row, qcol, max(0, metrics_row - 8), metrics_row
        )
        if not col_code:
            continue
        label = nearest_text(
            matrix[metrics_row - 1][qcol : qcol + 1] if metrics_row else []
        )
        if not label:
            label = nearest_text(
                matrix[r][qcol] for r in range(max(0, metrics_row - 8), metrics_row)
            )
        out.setdefault(
            ("", col_code),
            {
                "perimeter": perimeter,
                "template_code": template,
                "subtemplate_code": subtemplate,
                "row_code": "",
                "column_code": col_code,
                "qname": extract_qname(qname) or qname,
                "metric_label": extract_metric_label(qname) or label,
            },
        )
    return list(out.values())


def extract_records_from_matrix(
    matrix: List[List[str]], perimeter: str, template: str, subtemplate: str
) -> pl.DataFrame:
    if not matrix:
        return empty_metric_df()
    width = max(len(row) for row in matrix)
    matrix = [row + [""] * (width - len(row)) for row in matrix]
    rows_by_key: Dict[Tuple[str, str], Dict[str, str]] = {}
    metrics_cells = [
        (r, c)
        for r, row in enumerate(matrix)
        for c, cell in enumerate(row)
        if norm(cell).lower() == "metrics"
    ]
    for metrics_row, metrics_col in metrics_cells:
        for row in extract_row_oriented_rows(
            matrix, perimeter, template, subtemplate, metrics_row, metrics_col
        ):
            rows_by_key[(row["row_code"], row["column_code"])] = row
        for row in extract_column_oriented_rows(
            matrix, perimeter, template, subtemplate, metrics_row, metrics_col
        ):
            rows_by_key[(row["row_code"], row["column_code"])] = row
    if rows_by_key:
        return pl.DataFrame(list(rows_by_key.values()), schema=METRIC_COLS)

    row_codes = sorted(
        {
            first_coord(ROW_RE, cell)
            for row in matrix
            for cell in row
            if first_coord(ROW_RE, cell)
        }
    )
    col_codes = sorted(
        {
            first_coord(COL_RE, cell)
            for row in matrix
            for cell in row
            if first_coord(COL_RE, cell)
        }
    )
    fallback: List[Dict[str, str]] = []
    if row_codes:
        for row_code in row_codes:
            for col_code in col_codes or [""]:
                fallback.append(
                    {
                        "perimeter": perimeter,
                        "template_code": template,
                        "subtemplate_code": subtemplate,
                        "row_code": row_code,
                        "column_code": col_code,
                        "qname": "",
                        "metric_label": "",
                    }
                )
    else:
        for col_code in col_codes:
            fallback.append(
                {
                    "perimeter": perimeter,
                    "template_code": template,
                    "subtemplate_code": subtemplate,
                    "row_code": "",
                    "column_code": col_code,
                    "qname": "",
                    "metric_label": "",
                }
            )
    return pl.DataFrame(fallback, schema=METRIC_COLS) if fallback else empty_metric_df()

def extract_metrics(
    workbook: WorkbookCache,
    entries: Sequence[TocEntry],
    label: str = "Metrics",
) -> pl.DataFrame:

    sheets = workbook.sheet_names
    frames: List[pl.DataFrame] = []

    for entry in tqdm(
        entries,
        desc=label,
        unit="subtemplate",
        leave=True,
    ):
        found = empty_metric_df()

        for sheet in candidate_sheets(entry, sheets):
            window = subtemplate_window(
                workbook.matrix(sheet),
                entry.subtemplate_code,
            )

            found = extract_records_from_matrix(
                window,
                entry.perimeter,
                entry.template_code,
                entry.subtemplate_code,
            )

            if found.height:
                break

        if found.height:
            frames.append(found)
        else:
            LOG.debug(
                "No metrics found for %s/%s/%s",
                entry.perimeter,
                entry.template_code,
                entry.subtemplate_code,
            )

    if not frames:
        result = empty_metric_df()
    else:
        result = (
            pl.concat(frames, how="vertical")
            .unique(subset=KEY_COLS, keep="first")
        )

    LOG.info("Metrics extracted total: %d", result.height)
    return result

def build_metric_dataset(
    workbook: WorkbookCache,
    entries: Sequence[TocEntry],
    label: str,
) -> DpmDataset:

    selected_entries = list(entries)

    LOG.info(
        "Perimeters=%d Templates=%d Subtemplates=%d",
        len(extract_perimeters(selected_entries)),
        len(extract_templates(selected_entries)),
        len(extract_subtemplates(selected_entries)),
    )

    return DpmDataset(
        metrics=extract_metrics(
            workbook,
            selected_entries,
            label=label,
        ),
        entries=selected_entries,
    )


def to_delta_columns(df: pl.DataFrame) -> pl.DataFrame:
    return (
        df.select([pl.col(col).cast(pl.Utf8).fill_null("") for col in DELTA_COLS])
        if df.height
        else empty_delta_df()
    )


def metric_status_df(
    old: pl.DataFrame, new: pl.DataFrame
) -> Tuple[pl.DataFrame, pl.DataFrame, pl.DataFrame]:
    old_pref = old.rename({"qname": "QNameOld", "metric_label": "MetricLabelOld"})
    new_pref = new.rename({"qname": "QNameNew", "metric_label": "MetricLabelNew"})
    joined = old_pref.join(new_pref, on=KEY_COLS, how="inner")
    kept_modified = joined.with_columns(
        [
            pl.col("perimeter").alias("Perimeter"),
            pl.col("template_code").alias("TemplateCode"),
            pl.col("subtemplate_code").alias("SubTemplateCode"),
            pl.col("row_code").alias("RowCode"),
            pl.col("column_code").alias("ColumnCode"),
            pl.when(pl.col("QNameOld") != pl.col("QNameNew"))
            .then(pl.lit("Modified"))
            .otherwise(pl.lit("Kept"))
            .alias("Status"),
            metric_change_type_expr().alias("ChangeType"),
            pl.when(pl.col("QNameOld") != pl.col("QNameNew"))
            .then(
                pl.lit(
                    "QName changed for same template/subtemplate/row/column coordinates"
                )
            )
            .otherwise(pl.lit(""))
            .alias("Comments"),
        ]
    )
    old_only = old.join(new.select(KEY_COLS), on=KEY_COLS, how="anti")
    new_only = new.join(old.select(KEY_COLS), on=KEY_COLS, how="anti")
    return to_delta_columns(kept_modified), old_only, new_only


def qname_modified_df(
    old_only: pl.DataFrame, new_only: pl.DataFrame
) -> Tuple[pl.DataFrame, pl.DataFrame, pl.DataFrame]:
    old_q = old_only.filter(pl.col("qname") != "")
    new_q = new_only.filter(pl.col("qname") != "")
    if old_q.is_empty() or new_q.is_empty():
        return empty_delta_df(), old_only, new_only
    unique_old = (
        old_q.with_columns(pl.len().over(["perimeter", "qname"]).alias("_old_count"))
        .filter(pl.col("_old_count") == 1)
        .drop("_old_count")
    )
    unique_new = (
        new_q.with_columns(pl.len().over(["perimeter", "qname"]).alias("_new_count"))
        .filter(pl.col("_new_count") == 1)
        .drop("_new_count")
    )
    joined = unique_old.join(
        unique_new, on=["perimeter", "qname"], how="inner", suffix="_new"
    )
    if joined.is_empty():
        return empty_delta_df(), old_only, new_only
    changed = joined.filter(
        (pl.col("template_code") != pl.col("template_code_new"))
        | (pl.col("subtemplate_code") != pl.col("subtemplate_code_new"))
        | (pl.col("row_code") != pl.col("row_code_new"))
        | (pl.col("column_code") != pl.col("column_code_new"))
    )
    if changed.is_empty():
        return empty_delta_df(), old_only, new_only
    modified = changed.with_columns(
        [
            pl.col("perimeter").alias("Perimeter"),
            pl.col("template_code_new").alias("TemplateCode"),
            pl.col("subtemplate_code_new").alias("SubTemplateCode"),
            pl.col("row_code_new").alias("RowCode"),
            pl.col("column_code_new").alias("ColumnCode"),
            pl.col("qname").alias("QNameOld"),
            pl.col("qname").alias("QNameNew"),
            pl.col("metric_label").alias("MetricLabelOld"),
            pl.col("metric_label_new").alias("MetricLabelNew"),
            pl.lit("Modified").alias("Status"),
            metric_change_type_expr("row_code_new").alias("ChangeType"),
            pl.lit(
                "Same QName with changed template/subtemplate/row/column coordinates"
            ).alias("Comments"),
        ]
    )
    old_remove = changed.select(KEY_COLS)
    new_remove = changed.select(
        [
            pl.col("perimeter"),
            pl.col("template_code_new").alias("template_code"),
            pl.col("subtemplate_code_new").alias("subtemplate_code"),
            pl.col("row_code_new").alias("row_code"),
            pl.col("column_code_new").alias("column_code"),
        ]
    )
    remaining_old = old_only.join(old_remove, on=KEY_COLS, how="anti")
    remaining_new = new_only.join(new_remove, on=KEY_COLS, how="anti")
    return to_delta_columns(modified), remaining_old, remaining_new


def added_deleted_df(old_only: pl.DataFrame, new_only: pl.DataFrame) -> pl.DataFrame:
    frames: List[pl.DataFrame] = []
    if new_only.height:
        frames.append(
            to_delta_columns(
                new_only.with_columns(
                    [
                        pl.col("perimeter").alias("Perimeter"),
                        pl.col("template_code").alias("TemplateCode"),
                        pl.col("subtemplate_code").alias("SubTemplateCode"),
                        pl.col("row_code").alias("RowCode"),
                        pl.col("column_code").alias("ColumnCode"),
                        pl.lit("").alias("QNameOld"),
                        pl.col("qname").alias("QNameNew"),
                        pl.lit("").alias("MetricLabelOld"),
                        pl.col("metric_label").alias("MetricLabelNew"),
                        pl.lit("Added").alias("Status"),
                        (
                            pl.when(pl.col("row_code") != "")
                            .then(pl.lit("MetricRow"))
                            .otherwise(pl.lit("MetricColumn"))
                        ).alias("ChangeType"),
                        pl.lit("").alias("Comments"),
                    ]
                )
            )
        )
    if old_only.height:
        frames.append(
            to_delta_columns(
                old_only.with_columns(
                    [
                        pl.col("perimeter").alias("Perimeter"),
                        pl.col("template_code").alias("TemplateCode"),
                        pl.col("subtemplate_code").alias("SubTemplateCode"),
                        pl.col("row_code").alias("RowCode"),
                        pl.col("column_code").alias("ColumnCode"),
                        pl.col("qname").alias("QNameOld"),
                        pl.lit("").alias("QNameNew"),
                        pl.col("metric_label").alias("MetricLabelOld"),
                        pl.lit("").alias("MetricLabelNew"),
                        pl.lit("Deleted").alias("Status"),
                        (
                            pl.when(pl.col("row_code") != "")
                            .then(pl.lit("MetricRow"))
                            .otherwise(pl.lit("MetricColumn"))
                        ).alias("ChangeType"),
                        pl.lit("").alias("Comments"),
                    ]
                )
            )
        )
    return pl.concat(frames, how="vertical") if frames else empty_delta_df()


def structural_delta_df(old: DpmDataset, new: DpmDataset) -> pl.DataFrame:
    rows: List[Dict[str, str]] = []
    for perimeter, template in sorted(new.templates - old.templates):
        rows.append(
            {
                "Perimeter": perimeter,
                "TemplateCode": template,
                "SubTemplateCode": "",
                "RowCode": "",
                "ColumnCode": "",
                "QNameOld": "",
                "QNameNew": "",
                "MetricLabelOld": "",
                "MetricLabelNew": "",
                "Status": "Added",
                "ChangeType": "Template",
                "Comments": "Template present only in new version",
            }
        )
    for perimeter, template in sorted(old.templates - new.templates):
        rows.append(
            {
                "Perimeter": perimeter,
                "TemplateCode": template,
                "SubTemplateCode": "",
                "RowCode": "",
                "ColumnCode": "",
                "QNameOld": "",
                "QNameNew": "",
                "MetricLabelOld": "",
                "MetricLabelNew": "",
                "Status": "Deleted",
                "ChangeType": "Template",
                "Comments": "Template present only in old version",
            }
        )
    for perimeter, template, subtemplate in sorted(new.subtemplates - old.subtemplates):
        rows.append(
            {
                "Perimeter": perimeter,
                "TemplateCode": template,
                "SubTemplateCode": subtemplate,
                "RowCode": "",
                "ColumnCode": "",
                "QNameOld": "",
                "QNameNew": "",
                "MetricLabelOld": "",
                "MetricLabelNew": "",
                "Status": "Added",
                "ChangeType": "SubTemplate",
                "Comments": "Subtemplate present only in new version",
            }
        )
    for perimeter, template, subtemplate in sorted(old.subtemplates - new.subtemplates):
        rows.append(
            {
                "Perimeter": perimeter,
                "TemplateCode": template,
                "SubTemplateCode": subtemplate,
                "RowCode": "",
                "ColumnCode": "",
                "QNameOld": "",
                "QNameNew": "",
                "MetricLabelOld": "",
                "MetricLabelNew": "",
                "Status": "Deleted",
                "ChangeType": "SubTemplate",
                "Comments": "Subtemplate present only in old version",
            }
        )
    return pl.DataFrame(rows, schema=DELTA_COLS) if rows else empty_delta_df()


def compare_versions(old: DpmDataset, new: DpmDataset) -> pl.DataFrame:
    LOG.info("Delta step 1/5: coordinate comparison")
    kept_modified, old_only, new_only = metric_status_df(
        old.metrics,
        new.metrics,
    )

    LOG.info("Delta step 2/5: QName move detection")
    qname_modified, old_remaining, new_remaining = qname_modified_df(
        old_only,
        new_only,
    )

    LOG.info("Delta step 3/5: added/deleted metrics")
    added_deleted = added_deleted_df(
        old_remaining,
        new_remaining,
    )

    LOG.info("Delta step 4/5: structural changes")
    structural = structural_delta_df(old, new)

    LOG.info("Delta step 5/5: final assembly")
    frames = [
        df
        for df in [
            kept_modified,
            qname_modified,
            added_deleted,
            structural,
        ]
        if df.height
    ]

    return pl.concat(frames, how="vertical") if frames else empty_delta_df()

def generate_summary(delta: pl.DataFrame) -> Dict[str, float]:
    if delta.is_empty():
        return {
            k: 0.0
            for k in [
                "Metric Added",
                "Metric Added %",
                "Metric Deleted",
                "Metric Deleted %",
                "Metric Modified",
                "Metric Modified %",
                "Metric Kept",
                "Metric Kept %",
                "Template Added",
                "Template Added %",
                "Template Deleted",
                "Template Deleted %",
                "Subtemplate Added",
                "Subtemplate Added %",
                "Subtemplate Deleted",
                "Subtemplate Deleted %",
            ]
        }

    counts = {
        (r["ChangeType"], r["Status"]): r["len"]
        for r in delta.group_by(["ChangeType", "Status"]).len().to_dicts()
    }

    def metric_count(status: str) -> int:
        return counts.get(("MetricRow", status), 0) + counts.get(
            ("MetricColumn", status), 0
        )

    metric_added = metric_count("Added")
    metric_deleted = metric_count("Deleted")
    metric_modified = metric_count("Modified")
    metric_kept = metric_count("Kept")

    metric_total = metric_added + metric_deleted + metric_modified + metric_kept

    template_total = counts.get(("Template", "Added"), 0) + counts.get(
        ("Template", "Deleted"), 0
    )

    subtemplate_total = counts.get(("SubTemplate", "Added"), 0) + counts.get(
        ("SubTemplate", "Deleted"), 0
    )

    out: Dict[str, float] = {}

    out["Metric Added"] = metric_added
    out["Metric Added %"] = metric_added / metric_total if metric_total else 0.0

    out["Metric Deleted"] = metric_deleted
    out["Metric Deleted %"] = metric_deleted / metric_total if metric_total else 0.0

    out["Metric Modified"] = metric_modified
    out["Metric Modified %"] = metric_modified / metric_total if metric_total else 0.0

    out["Metric Kept"] = metric_kept
    out["Metric Kept %"] = metric_kept / metric_total if metric_total else 0.0

    template_added = counts.get(("Template", "Added"), 0)
    template_deleted = counts.get(("Template", "Deleted"), 0)

    out["Template Added"] = template_added
    out["Template Added %"] = template_added / template_total if template_total else 0.0

    out["Template Deleted"] = template_deleted
    out["Template Deleted %"] = (
        template_deleted / template_total if template_total else 0.0
    )

    subtemplate_added = counts.get(("SubTemplate", "Added"), 0)
    subtemplate_deleted = counts.get(("SubTemplate", "Deleted"), 0)

    out["Subtemplate Added"] = subtemplate_added
    out["Subtemplate Added %"] = (
        subtemplate_added / subtemplate_total if subtemplate_total else 0.0
    )

    out["Subtemplate Deleted"] = subtemplate_deleted
    out["Subtemplate Deleted %"] = (
        subtemplate_deleted / subtemplate_total if subtemplate_total else 0.0
    )

    return out


def safe_sheet_name(name: str, used: Set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", name)[:31] or "perimeter"
    candidate = base
    i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    used.add(candidate)
    return candidate


def apply_formatting(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
    for row in ws.iter_rows():
        status = row[9].value if row[0].row > 1 and len(row) >= 10 else None
        fill = {
            "Added": ADDED_FILL,
            "Deleted": DELETED_FILL,
            "Modified": MODIFIED_FILL,
            "Kept": KEPT_FILL,
        }.get(status, KEPT_FILL)
        for cell in row:
            cell.border = BORDER
            if cell.row > 1:
                cell.fill = fill
    for col_idx, col in enumerate(ws.columns, start=1):
        width = min(
            80,
            max(
                10,
                max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col
                )
                + 2,
            ),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def generate_delta_workbook(delta: pl.DataFrame, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary = generate_summary(delta)
    ws.append(list(summary.keys()))
    ws.append([summary[key] for key in summary])
    for cell in ws[2]:
        if isinstance(cell.value, float):
            cell.number_format = "0.00%"
    apply_formatting(ws)

    used = {"Summary"}
    if not delta.is_empty():
        for perimeter in tqdm(sorted(delta.get_column("Perimeter").unique().to_list()), desc="Writing delta workbook", unit="sheet"):
            ws = wb.create_sheet(safe_sheet_name(str(perimeter), used))
            ws.append(DELTA_COLS)
            subset = delta.filter(pl.col("Perimeter") == perimeter).sort(
                [
                    "ChangeType",
                    "TemplateCode",
                    "SubTemplateCode",
                    "RowCode",
                    "ColumnCode",
                    "Status",
                ]
            )
            for row in subset.iter_rows(named=False):
                ws.append(list(row))
            apply_formatting(ws)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare two EIOPA DPM Excel workbooks and create a delta workbook"
    )
    parser.add_argument("old_version", type=Path)
    parser.add_argument("new_version", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument(
        "--no-interactive-perimeters",
        action="store_true",
        help="Do not show perimeter selection; include all perimeters",
    )
    parser.add_argument(
        "--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"]
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(name)s - %(message)s",
    )

    old_book = WorkbookCache.open(args.old_version)
    new_book = WorkbookCache.open(args.new_version)

    try:
        if not args.old_version.exists():
            raise FileNotFoundError(args.old_version)
        if not args.new_version.exists():
            raise FileNotFoundError(args.new_version)
        old_entries_all = extract_toc(old_book)
        new_entries_all = extract_toc(new_book)
        if args.no_interactive_perimeters:
            selected_perimeters = {
                p.lower()
                for p in sorted(
                    extract_perimeters(old_entries_all)
                    | extract_perimeters(new_entries_all)
                )
            }
            LOG.info("Selected perimeters: all")
        else:
            selected_perimeters = select_perimeters_interactively(
                old_entries_all, new_entries_all
            )
            LOG.info("Selected perimeters: %s", sorted(selected_perimeters))
        old_entries = filter_entries_by_perimeter(old_entries_all, selected_perimeters)
        new_entries = filter_entries_by_perimeter(new_entries_all, selected_perimeters)
        old_dataset = build_metric_dataset(
            old_book,
            old_entries,
            label="Extracting OLD metrics",
        )

        new_dataset = build_metric_dataset(
            new_book,
            new_entries,
            label="Extracting NEW metrics",
        )
        delta = compare_versions(old_dataset, new_dataset)
        generate_delta_workbook(delta, args.output)
        LOG.info("Delta workbook generated: %s", args.output)
        return 0
    except KeyboardInterrupt:
        LOG.error("Cancelled by user")
        return 130
    except Exception:
        LOG.exception("DPM delta generation failed")
        return 1
    finally:
        old_book.close()
        new_book.close()


if __name__ == "__main__":
    raise SystemExit(main())
