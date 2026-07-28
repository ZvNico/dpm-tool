#!/usr/bin/env python3
from __future__ import annotations

import argparse
import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import polars as pl
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: polars. Install with: pip install polars"
    ) from exc

from openpyxl import load_workbook

LOG = logging.getLogger("xbrl_apply_delta_regex_polars")

DELTA_HEADERS = [
    "Perimeter",
    "TemplateCode",
    "SubTemplateCode",
    "RowCode",
    "ColumnCode",
    "QNameOld",
    "QNameNew",
    "MetricLabelOld",
    "MetricLabelNew",
    "Status",
    "ChangeType",
    "Comments",
]

FACT_SCHEMA = {
    "fact_index": pl.Int64,
    "start": pl.Int64,
    "end": pl.Int64,
    "qname": pl.Utf8,
    "attributes": pl.Utf8,
    "value": pl.Utf8,
}

DELTA_SCHEMA = {header: pl.Utf8 for header in DELTA_HEADERS}

# Captures normal metric facts, including multiline values.
# Example: <s2md_met:ei2751 contextRef="C1">s2c_MC:x0</s2md_met:ei2751>
METRIC_FACT_RE = re.compile(
    rb"<(?P<qname>s2md_met:[A-Za-z_][A-Za-z0-9_.-]*)\b(?P<attrs>[^>]*)>"
    rb"(?P<value>.*?)"
    rb"</(?P=qname)>",
    re.DOTALL,
)

# Optional support for self-closing metric facts.
SELF_CLOSING_METRIC_FACT_RE = re.compile(
    rb"<(?P<qname>s2md_met:[A-Za-z_][A-Za-z0-9_.-]*)\b(?P<attrs>[^>]*)/>",
    re.DOTALL,
)

SCHEMA_REF_RE = re.compile(
    rb"<(?P<tag>(?:[A-Za-z_][A-Za-z0-9_.-]*:)?schemaRef)\b[^>]*\b(?:[A-Za-z_][A-Za-z0-9_.-]*:)?href\s*=\s*(['\"])(?P<href>.*?)\2[^>]*/?>",
    re.DOTALL,
)


@dataclass(frozen=True)
class ApplyStats:
    perimeter: str
    facts_before: int
    facts_after: int
    deleted_facts: int
    renamed_facts: int
    deleted_qnames: int
    modified_qnames: int


def norm(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fast byte/regex XBRL updater: delete and rename s2md_met facts using Delta_DPM.xlsx."
    )
    parser.add_argument(
        "delta_workbook", type=Path, help="Delta workbook produced by dpm_delta.py"
    )
    parser.add_argument("input_xbrl", type=Path, help="Input XBRL instance")
    parser.add_argument(
        "output_xbrl", type=Path, help="Output XBRL instance. Must differ from input."
    )
    parser.add_argument(
        "--perimeter", help="Override perimeter detected from schemaRef, e.g. qrs"
    )
    parser.add_argument("--dry-run", action="store_true", help="Do not write output")
    parser.add_argument(
        "--facts-parquet",
        type=Path,
        help="Optional parquet dump of flattened metric facts after delta flags",
    )
    parser.add_argument(
        "--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"]
    )
    return parser.parse_args()


def detect_perimeter_from_xml_bytes(xml: bytes) -> str:
    candidates: List[str] = []
    for match in SCHEMA_REF_RE.finditer(xml):
        href = match.group("href").decode("utf-8", errors="replace")
        mod_match = re.search(r"/mod/([a-z0-9_-]+)\.xsd(?:$|[?#])", href, re.I)
        if mod_match:
            candidates.append(mod_match.group(1).lower())
            continue
        filename = href.rstrip("/").split("/")[-1]
        if filename.lower().endswith(".xsd"):
            candidates.append(filename[:-4].lower())
    if not candidates:
        raise ValueError(
            "Unable to detect perimeter from schemaRef href. Use --perimeter."
        )
    unique = list(dict.fromkeys(candidates))
    if len(unique) > 1:
        LOG.warning(
            "Multiple perimeter candidates found in schemaRef hrefs: %s. Using %s",
            unique,
            unique[0],
        )
    return unique[0]


def worksheet_rows(ws) -> Iterable[Dict[str, str]]:
    iterator = ws.iter_rows(values_only=True)
    try:
        header = [norm(value) for value in next(iterator)]
    except StopIteration:
        return
    index = {name: pos for pos, name in enumerate(header)}
    if not all(name in index for name in DELTA_HEADERS):
        return
    for raw in iterator:
        if not raw or all(norm(value) == "" for value in raw):
            continue
        yield {
            name: norm(raw[index[name]]) if index[name] < len(raw) else ""
            for name in DELTA_HEADERS
        }


def load_delta_df(delta_workbook: Path, perimeter: str) -> pl.DataFrame:
    workbook = load_workbook(delta_workbook, read_only=True, data_only=True)
    try:
        sheet_name = next(
            (name for name in workbook.sheetnames if name.lower() == perimeter.lower()),
            None,
        )
        worksheets = (
            [workbook[sheet_name]]
            if sheet_name
            else [
                workbook[name]
                for name in workbook.sheetnames
                if name.lower() != "summary"
            ]
        )
        rows: List[Dict[str, str]] = []
        for ws in worksheets:
            for row in worksheet_rows(ws):
                if (
                    row["Perimeter"].lower() == perimeter.lower()
                    and row["ChangeType"].lower() in {"metricrow", "metriccolumn"}
                ):
                    rows.append(row)
        if not rows:
            return pl.DataFrame(schema=DELTA_SCHEMA)
        return pl.DataFrame(rows, schema=DELTA_SCHEMA)
    finally:
        workbook.close()


def build_delta_maps(delta: pl.DataFrame) -> Tuple[pl.DataFrame, pl.DataFrame]:
    if delta.is_empty():
        return pl.DataFrame(schema={"qname": pl.Utf8}), pl.DataFrame(
            schema={"qname": pl.Utf8, "qname_new": pl.Utf8}
        )

    deleted = (
        delta.filter(
            (pl.col("Status") == "Deleted")
            & ((pl.col("QNameOld") != "") | (pl.col("QNameNew") != ""))
        )
        .with_columns(
            pl.when(pl.col("QNameOld") != "")
            .then(pl.col("QNameOld"))
            .otherwise(pl.col("QNameNew"))
            .alias("qname")
        )
        .select("qname")
        .unique()
    )

    modified = (
        delta.filter(
            (pl.col("Status") == "Modified")
            & (pl.col("QNameOld") != "")
            & (pl.col("QNameNew") != "")
            & (pl.col("QNameOld") != pl.col("QNameNew"))
        )
        .select(
            [pl.col("QNameOld").alias("qname"), pl.col("QNameNew").alias("qname_new")]
        )
        .unique(subset=["qname"], keep="first")
    )

    return deleted, modified


def flatten_metric_facts(xml: bytes) -> pl.DataFrame:
    rows: List[Dict[str, object]] = []
    fact_index = 0

    for match in METRIC_FACT_RE.finditer(xml):
        rows.append(
            {
                "fact_index": fact_index,
                "start": match.start(),
                "end": match.end(),
                "qname": match.group("qname").decode("utf-8", errors="replace"),
                "attributes": match.group("attrs").decode("utf-8", errors="replace"),
                "value": match.group("value").decode("utf-8", errors="replace"),
            }
        )
        fact_index += 1

    # Only add self-closing facts that do not overlap normal facts.
    occupied = [(int(row["start"]), int(row["end"])) for row in rows]
    for match in SELF_CLOSING_METRIC_FACT_RE.finditer(xml):
        start, end = match.start(), match.end()
        if any(start >= lo and end <= hi for lo, hi in occupied):
            continue
        rows.append(
            {
                "fact_index": fact_index,
                "start": start,
                "end": end,
                "qname": match.group("qname").decode("utf-8", errors="replace"),
                "attributes": match.group("attrs").decode("utf-8", errors="replace"),
                "value": "",
            }
        )
        fact_index += 1

    if not rows:
        return pl.DataFrame(schema=FACT_SCHEMA)
    return pl.DataFrame(rows, schema=FACT_SCHEMA).sort("start")


def flag_facts_with_delta(
    facts: pl.DataFrame, deleted: pl.DataFrame, modified: pl.DataFrame
) -> pl.DataFrame:
    if facts.is_empty():
        return facts.with_columns(
            [
                pl.lit(False).alias("delete_fact"),
                pl.lit(False).alias("rename_fact"),
                pl.col("qname").alias("qname_out"),
            ]
        )

    result = facts
    if deleted.is_empty():
        result = result.with_columns(pl.lit(False).alias("delete_fact"))
    else:
        result = result.join(
            deleted.with_columns(pl.lit(True).alias("delete_fact")),
            on="qname",
            how="left",
        ).with_columns(pl.col("delete_fact").fill_null(False))

    if modified.is_empty():
        result = result.with_columns(pl.lit(None).cast(pl.Utf8).alias("qname_new"))
    else:
        result = result.join(modified, on="qname", how="left")

    return result.with_columns(
        [
            pl.col("qname_new").is_not_null().alias("rename_fact"),
            pl.when(pl.col("qname_new").is_not_null())
            .then(pl.col("qname_new"))
            .otherwise(pl.col("qname"))
            .alias("qname_out"),
        ]
    )


def rename_fact_bytes(segment: bytes, old_qname: str, new_qname: str) -> bytes:
    old_b = old_qname.encode("utf-8")
    new_b = new_qname.encode("utf-8")
    pattern = re.compile(rb"(<\/?)" + re.escape(old_b) + rb"(\b)")
    return pattern.sub(
        lambda match: match.group(1) + new_b + match.group(2), segment, count=2
    )


def expand_deleted_fact_span(xml: bytes, start: int, end: int) -> Tuple[int, int]:
    # Delete the whole physical line of the metric fact to avoid leaving blank lines.
    # If the tag is indented, remove the indentation before the opening tag too.
    line_start = xml.rfind(b"\n", 0, start) + 1
    if line_start < 0:
        line_start = 0
    if xml[line_start:start].strip() == b"":
        start = line_start

    # Consume one following line break after the closing tag, but keep indentation of the next line.
    if end < len(xml) and xml[end : end + 2] == b"\r\n":
        end += 2
    elif end < len(xml) and xml[end : end + 1] in {b"\n", b"\r"}:
        end += 1
    return start, end


def rebuild_xml_bytes(xml: bytes, facts: pl.DataFrame) -> bytes:
    changed = (
        facts.filter((pl.col("delete_fact") == True) | (pl.col("rename_fact") == True))
        .select(["start", "end", "qname", "qname_out", "delete_fact", "rename_fact"])
        .sort("start")
        .to_dicts()
    )
    if not changed:
        return xml

    parts: List[bytes] = []
    cursor = 0
    for row in changed:
        start = int(row["start"])
        end = int(row["end"])
        if row["delete_fact"]:
            start, end = expand_deleted_fact_span(xml, start, end)
            if end <= cursor:
                continue
            start = max(start, cursor)
            parts.append(xml[cursor:start])
            cursor = end
            continue
        if start < cursor:
            raise ValueError("Overlapping metric fact matches detected")
        parts.append(xml[cursor:start])
        if row["rename_fact"]:
            parts.append(
                rename_fact_bytes(
                    xml[start:end], str(row["qname"]), str(row["qname_out"])
                )
            )
        else:
            parts.append(xml[start:end])
        cursor = end
    parts.append(xml[cursor:])
    return b"".join(parts)


def apply_delta(
    delta_workbook: Path,
    input_xbrl: Path,
    output_xbrl: Path,
    perimeter_override: Optional[str],
    dry_run: bool,
    facts_parquet: Optional[Path],
) -> ApplyStats:
    if output_xbrl.resolve() == input_xbrl.resolve():
        raise ValueError("Output path must be different from input XBRL path")

    xml = input_xbrl.read_bytes()
    perimeter = (
        perimeter_override.lower()
        if perimeter_override
        else detect_perimeter_from_xml_bytes(xml)
    )
    LOG.info("Detected perimeter: %s", perimeter)

    delta = load_delta_df(delta_workbook, perimeter)
    deleted, modified = build_delta_maps(delta)
    LOG.info("Delta qnames: deleted=%d modified=%d", deleted.height, modified.height)

    facts = flatten_metric_facts(xml)
    LOG.info("Flattened metric facts: %d", facts.height)

    flagged = flag_facts_with_delta(facts, deleted, modified)
    deleted_facts = flagged.filter(pl.col("delete_fact") == True).height
    renamed_facts = flagged.filter(
        (pl.col("delete_fact") == False) & (pl.col("rename_fact") == True)
    ).height

    if facts_parquet:
        facts_parquet.parent.mkdir(parents=True, exist_ok=True)
        flagged.write_parquet(facts_parquet)
        LOG.info("Fact dataframe written to: %s", facts_parquet)

    if not dry_run:
        output_xml = rebuild_xml_bytes(xml, flagged)
        output_xbrl.parent.mkdir(parents=True, exist_ok=True)
        output_xbrl.write_bytes(output_xml)
        LOG.info("Updated XBRL written to: %s", output_xbrl)
    else:
        LOG.info("Dry run enabled; output file not written")

    return ApplyStats(
        perimeter=perimeter,
        facts_before=facts.height,
        facts_after=facts.height - deleted_facts,
        deleted_facts=deleted_facts,
        renamed_facts=renamed_facts,
        deleted_qnames=deleted.height,
        modified_qnames=modified.height,
    )


def main() -> int:
    args = parse_args()
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(name)s - %(message)s",
    )
    try:
        if not args.delta_workbook.exists():
            raise FileNotFoundError(args.delta_workbook)
        if not args.input_xbrl.exists():
            raise FileNotFoundError(args.input_xbrl)

        stats = apply_delta(
            delta_workbook=args.delta_workbook,
            input_xbrl=args.input_xbrl,
            output_xbrl=args.output_xbrl,
            perimeter_override=args.perimeter,
            dry_run=args.dry_run,
            facts_parquet=args.facts_parquet,
        )
        LOG.info(
            "Result: perimeter=%s facts_before=%d facts_after=%d deleted_facts=%d renamed_facts=%d deleted_qnames=%d modified_qnames=%d",
            stats.perimeter,
            stats.facts_before,
            stats.facts_after,
            stats.deleted_facts,
            stats.renamed_facts,
            stats.deleted_qnames,
            stats.modified_qnames,
        )
        return 0
    except Exception:
        LOG.exception("XBRL delta application failed")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
