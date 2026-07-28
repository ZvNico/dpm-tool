from __future__ import annotations

import re
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

from dpm._constants import (
    ADDED_FILL,
    BORDER,
    DELTA_COLS,
    DELETED_FILL,
    HEADER_FILL,
    KEPT_FILL,
    MODIFIED_FILL,
)

_STATUS_COL_IDX = DELTA_COLS.index("Status")
from dpm.delta import generate_summary

_STATUS_FILL = {
    "Added": ADDED_FILL,
    "Deleted": DELETED_FILL,
    "Modified": MODIFIED_FILL,
    "Kept": KEPT_FILL,
}


def _safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", name)[:31] or "perimeter"
    candidate = base
    i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    used.add(candidate)
    return candidate


def _apply_formatting(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
    for row in ws.iter_rows():
        status = row[_STATUS_COL_IDX].value if row[0].row > 1 and len(row) > _STATUS_COL_IDX else None
        fill = _STATUS_FILL.get(status, KEPT_FILL)
        for cell in row:
            cell.border = BORDER
            if cell.row > 1:
                cell.fill = fill
    for col_idx, col in enumerate(ws.columns, start=1):
        width = min(
            80,
            max(10, max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def generate_delta_workbook(delta: pl.DataFrame, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary = generate_summary(delta)
    ws.append(list(summary.keys()))
    ws.append(list(summary.values()))
    for cell in ws[2]:
        if isinstance(cell.value, float):
            cell.number_format = "0.00%"
    _apply_formatting(ws)

    used: set[str] = {"Summary"}
    if not delta.is_empty():
        for perimeter in tqdm(
            sorted(delta.get_column("Perimeter").unique().to_list()),
            desc="Writing delta workbook",
            unit="sheet",
        ):
            ws = wb.create_sheet(_safe_sheet_name(str(perimeter), used))
            ws.append(DELTA_COLS)
            subset = delta.filter(pl.col("Perimeter") == perimeter).sort(
                ["ChangeType", "TemplateCode", "SubTemplateCode", "RowCode", "ColumnCode", "Status"]
            )
            for row in subset.iter_rows(named=False):
                ws.append(list(row))
            _apply_formatting(ws)
    wb.save(output_path)
