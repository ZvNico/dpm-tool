from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import TypedDict

import polars as pl
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from dpm._text import norm


@dataclass(frozen=True)
class TocEntry:
    perimeter: str
    template_code: str
    subtemplate_code: str
    sheet_name: str | None = None


@dataclass(frozen=True)
class DpmDataset:
    metrics: pl.DataFrame
    entries: list[TocEntry]
    # Full metric catalogue (metric_code, metric_label) for the metrics delta.
    metric_catalog: pl.DataFrame = field(default_factory=pl.DataFrame)
    # Flat dimensions × members frame for the dimensions delta.
    dimension_members: pl.DataFrame = field(default_factory=pl.DataFrame)

    @property
    def templates(self) -> set[tuple[str, str]]:
        return {(e.perimeter, e.template_code) for e in self.entries}

    @property
    def subtemplates(self) -> set[tuple[str, str, str]]:
        return {
            (e.perimeter, e.template_code, e.subtemplate_code) for e in self.entries
        }


@dataclass(frozen=True)
class DeltaResult:
    """The three-section delta between two DPM versions.

    Each frame follows the canonical snake_case schema in ``dpm.delta_schema``:
    ``structure`` the per-perimeter fact/structural delta (``DELTA_STRUCTURE_COLS``),
    ``metrics`` the metric-catalogue delta (``DELTA_METRIC_COLS``), and
    ``dimensions`` the dimensions/members delta (``DELTA_DIMENSION_COLS``).
    """

    structure: pl.DataFrame
    metrics: pl.DataFrame
    dimensions: pl.DataFrame


class TemplateRow(TypedDict):
    template_code: str
    template_label: str


class SubtemplateRow(TypedDict):
    subtemplate_code: str
    template_code: str
    subtemplate_label: str
    subtemplate_type: str


class PerimeterRow(TypedDict):
    perimeter_code: str


class PerimeterTemplateRow(TypedDict):
    perimeter_code: str
    template_code: str


class MetricRow(TypedDict):
    metric_code: str
    metric_label: str


class FactRow(TypedDict):
    subtemplate_code: str
    row_code: str
    column_code: str
    row_label: str
    column_label: str
    metric_code: str


class DimensionRow(TypedDict):
    dimension_code: str
    dimension_label: str


class DimensionMemberRow(TypedDict):
    member_code: str
    dimension_code: str
    member_label: str


class FactDimensionRow(TypedDict):
    subtemplate_code: str
    row_code: str
    column_code: str
    dimension_code: str
    member_code: str


@dataclass
class WorkbookCache:
    path: Path
    wb: Workbook
    sheet_names: list[str]
    _matrix_cache: dict[str, list[list[str]]] = field(
        default_factory=dict, repr=False, init=False
    )
    _crossed_cache: dict[str, frozenset[tuple[int, int]]] = field(
        default_factory=dict, repr=False, init=False
    )

    @classmethod
    def open(cls, path: Path) -> WorkbookCache:
        wb = load_workbook(path, read_only=True, data_only=True)
        return cls(path=path, wb=wb, sheet_names=list(wb.sheetnames))

    def close(self) -> None:
        self.wb.close()

    def _ensure_parsed(self, sheet_name: str) -> None:
        if sheet_name in self._matrix_cache:
            return
        ws = self.wb[sheet_name]
        rows: list[list[str]] = []
        crossed: set[tuple[int, int]] = set()
        for r_idx, row in enumerate(ws.iter_rows()):
            row_vals: list[str] = []
            for cell in row:
                row_vals.append(norm(cell.value))
                try:
                    if cell.fill and cell.fill.fill_type == "solid":
                        col_idx = getattr(cell, "column", None)
                        if col_idx is not None:
                            crossed.add((r_idx, col_idx - 1))
                except (AttributeError, TypeError):
                    pass
            rows.append(row_vals)
        self._matrix_cache[sheet_name] = rows
        self._crossed_cache[sheet_name] = frozenset(crossed)

    def matrix(self, sheet_name: str) -> list[list[str]]:
        self._ensure_parsed(sheet_name)
        return self._matrix_cache[sheet_name]

    def crossed_cells(self, sheet_name: str) -> frozenset[tuple[int, int]]:
        """Return 0-indexed (row, col) positions where cells have a solid fill (invalid data cells)."""
        self._ensure_parsed(sheet_name)
        return self._crossed_cache.get(sheet_name, frozenset())


@dataclass(frozen=True)
class ApplyStats:
    perimeter: str
    facts_before: int
    facts_after: int
    deleted_facts: int
    renamed_facts: int
    deleted_qnames: int
    modified_qnames: int
