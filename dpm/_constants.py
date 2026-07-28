from __future__ import annotations

import re

from openpyxl.styles import Border, PatternFill, Side

ROW_RE = re.compile(r"\b[A-Z]{0,3}R\d{3,6}\b", re.I)
COL_RE = re.compile(r"\b[A-Z]{0,3}C\d{3,6}\b", re.I)
CODE_RE = re.compile(r"\b[A-Z]{1,3}\.(?:[0-9]{2}\.){1,6}[0-9]{2}\b", re.I)
QNAME_RE = re.compile(
    r"(^\{[^}]+\}.+)|(^[A-Za-z_][\w.-]*:[A-Za-z_][\w.-]*$)"
    r"|(^[A-Za-z_][\w.-]*\.[A-Za-z_][\w.-]*(?:\.[A-Za-z_][\w.-]*)+$)"
)
QNAME_TOKEN_RE = re.compile(r"\b(?:s2md_met|[A-Za-z_][\w.-]*):[A-Za-z_][\w.-]*\b")
METRIC_LABEL_RE = re.compile(r"Metric:\s*(.*?)(?:\)\s*(?:\[|$)|$)")
GENERIC_TOC_CODES = frozenset({"T99", "T.99", "T.99.99", "TOC", "TABLE OF CONTENTS"})

METRIC_COLS = [
    "perimeter",
    "template_code",
    "subtemplate_code",
    "row_code",
    "column_code",
    "qname",
    "metric_label",
    "row_label",
    "column_label",
]
KEY_COLS = ["perimeter", "template_code", "subtemplate_code", "row_code", "column_code"]
IDENTITY_COLS = ["template_code", "subtemplate_code", "row_code", "column_code"]
DELTA_COLS = [
    "Perimeter",
    "TemplateCode",
    "SubTemplateCode",
    "RowCode",
    "ColumnCode",
    "QNameOld",
    "QNameNew",
    "MetricLabelOld",
    "MetricLabelNew",
    "RowLabelOld",
    "RowLabelNew",
    "ColumnLabelOld",
    "ColumnLabelNew",
    "Status",
    "ChangeType",
    "Comments",
]

ADDED_FILL = PatternFill("solid", fgColor="C6EFCE")
DELETED_FILL = PatternFill("solid", fgColor="FFC7CE")
MODIFIED_FILL = PatternFill("solid", fgColor="FCE4D6")
KEPT_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
_THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
